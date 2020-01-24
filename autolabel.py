# coding=utf-8
import argparse
import io
import os
import shutil
import time
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook


def load_defaults(table_path):
    """
    Читает из таблицы настройки для обработки изображений
    :param table_path: str, путь до таблицы с настройками
    :return: max_res_x: int, максимальное разрешение по x
    :return: max_res_y: int, максимальное разрешение по y
    :return: max_size: float, максимальный размер в Мб
    :return: sample_text: str, стандартный текст (при отсутствии такового в таблице)
    :return: opacity: int, прозрачность фона этикетки
    :return: font_size: int, размер шрифта
    """
    wb = load_workbook(filename=table_path)
    ws = wb['example']
    max_res_x = ws['D2'].value
    max_res_y = ws['E2'].value
    max_size = ws['F2'].value
    sample_text = ws['C2'].value
    opacity = ws['H2'].value
    font_size = ws['G2'].value
    return max_res_x, max_res_y, max_size, sample_text, opacity, font_size


def load_rows_from_xlsx(table_path, sample_text=None):
    """
    Загружает пути и тексты для обработки фотографий из таблицы
    :param table_path: str, путь до таблицы
    :param sample_text: str, стандартный текст при отсутствии такового в таблице
    :return: dict, словарь "{путь}": "{текст}", или None если в таблице не было найдено путей до файлов
    """
    wb = load_workbook(filename=table_path)
    ws = wb['example']
    file_dict = {}
    index = 3
    while ws['B' + str(index)].value is not None:
        text = ws['C' + str(index)].value
        if text is None:
            text = sample_text
        file_dict[Path(ws['B' + str(index)].value)] = text
        index += 1
    if file_dict != {}:
        return file_dict
    else:
        return None


def write_files_in_xlsx(table_path):
    """
    Считывает файлы из папки по пути, записанному в определенную ячейку
    таблицы по пути table_path, и записывает их пути обратно в таблицу
    :param table_path: str, путь до таблицы
    """
    work_book = load_workbook(table_path)
    work_sheet = work_book['example']
    file_path = work_sheet['B2'].value
    file_list = [entry for entry in os.scandir(file_path) if entry.is_file()
                 and (entry.path.lower().endswith('jpeg') or entry.path.lower().endswith('jpg')
                      or entry.path.lower().endswith('png'))]
    for file in file_list:
        work_sheet.cell(row=file_list.index(file) + 3, column=2).value = file.path
    work_book.save(table_path)


def null_one(number):
    """
    Возвращает 0 при аргументе равном 1 и наоборот
    :param number: int
    :return: int
    """
    if number == 0:
        return 1
    if number == 1:
        return 0


def process_image(path, text, output_folder, max_res_x=3000, max_res_y=2250,
                  max_size=2.0, opacity=90, font_size=50, corner=(1, 1)):
    """
    Считывает изображение по пути path, сохраняет его в выходную папку с этикеткой нанесенной поверх изображения
    :param path: str, путь до исходного изображения
    :param text: str, текст этикетки
    :param output_folder: str, папка с выходными данными
    :param max_res_x: int, максимальное разрешение по x
    :param max_res_y: int, максимальное разрешение по y
    :param max_size: float, максимальный размер в Мб
    :param opacity: int, прозрачность фона этикетки
    :param font_size: int, размер шрифта
    :param corner: tuple, угол для размещения этикетки на изображении
    """
    image = Image.open(path)
    # check x asis
    if image.size[0] > max_res_x:
        image = image.resize((max_res_x, image.size[1] * max_res_x // image.size[0]), Image.ANTIALIAS)

    # check y axis
    if image.size[0] > max_res_y:
        image = image.resize((image.size[0] * max_res_y // image.size[1], max_res_y), Image.ANTIALIAS)

    # check size
    q = 90
    while check_size(image, quality=q) > max_size * 1024 * 1024:
        q -= 5

    # add rectangle
    draw = ImageDraw.Draw(image, 'RGBA')
    font = ImageFont.truetype(str(Path(os.path.abspath(__file__)).parent / 'Lora-Regular.ttf'), font_size)
    text_size = draw.textsize(text, font)
    rectangle_coords = (((image.size[0] - text_size[0] - font_size) * corner[0],
                         (image.size[1] - text_size[1] - font_size) * corner[1]),
                        ((image.size[0] - 1) * corner[0] + (text_size[0] + font_size) * null_one(corner[0]),
                         (image.size[1] - 1) * corner[1] + (text_size[1] + font_size) * null_one(corner[1])))
    draw.rectangle(rectangle_coords, fill=(255, 255, 255, 255 * opacity // 100), outline=(0, 0, 0), width=4)
    # draw text
    draw.text((rectangle_coords[0][0] + font_size // 2, rectangle_coords[0][1] + font_size // 2), text, font=font,
              fill=(0, 0, 0, 250))

    # output file
    output_file = output_folder / (path.name.rsplit('.', 1)[0] + '_edit.jpg')
    print('Файл {}, вес {} Мб, JPEG quality {}, размер {}x{}'.
          format(output_file.name, round(check_size(image, quality=q) / 1024 / 1024, 2), q, image.size[0],
                 image.size[1]))
    image.save(output_file, format='jpeg', quality=q)


def check_size(image, quality=90):
    """
    Проверка размера изображения на выходе
    :param image: обьект класса PIL.Image
    :param quality: int, JPEG quality
    :return: int, размер изображения в байтах
    """
    out = io.BytesIO()
    image.save(out, format='jpeg', quality=quality)
    return out.tell()


def main(args):
    max_res_x, max_res_y, max_size, sample_text, opacity, font_size = load_defaults(args.table_path)
    if args.populating_mode:
        write_files_in_xlsx(args.table_path)
    if not args.disable_labeling:
        file_dict = load_rows_from_xlsx(args.table_path, sample_text=sample_text)
        if file_dict is None:
            exit('В таблице нет сведений о файлах для обработки!')
        # delete output folder
        output_folder = list(file_dict.keys())[0].parent / 'output'
        if output_folder.exists() and output_folder.is_dir():
            shutil.rmtree(output_folder)
        # create output folder
        print('Папка результатов: {}'.format(output_folder))
        time.sleep(0.5)  # stupid workaround for WinError 5
        Path(output_folder).mkdir(parents=True, exist_ok=True)
        # process files
        for key in file_dict:
            process_image(key, file_dict[key], output_folder,
                          max_res_x=max_res_x, max_res_y=max_res_y, max_size=max_size,
                          opacity=opacity, font_size=font_size)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Python-скрипт для этикетирования фотографий')
    parser.add_argument('-p', '--populate-table',
                        dest='populating_mode',
                        action='store_const',
                        const=True,
                        default=False,
                        required=False,
                        help='Включить заполнение таблицы, по умолчанию выключено')
    parser.add_argument('-dl', '--disable_labeling',
                        dest='disable_labeling',
                        action='store_const',
                        const=True,
                        default=False,
                        required=False,
                        help='Выключить обработку изображений')
    parser.add_argument('-t', '--table',
                        dest='table_path',
                        type=str,
                        default='example.xlsx',
                        required=False,
                        help='Путь до таблицы, с которой производит работу скрипт')
    args = parser.parse_args()
    main(args)
